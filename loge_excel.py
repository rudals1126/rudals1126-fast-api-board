import pandas as pd  
# → pandas 라이브러리를 불러옴. 엑셀 파일을 읽고(DataFrame으로) 쓰는 기능에 사용
from openpyxl import load_workbook  
# → openpyxl은 엑셀 파일의 시트(워크북)를 직접 다루는 라이브러리임.
#   특히 시트가 존재하는지 확인할 때 pandas보다 빠르고 안정적이라 사용
from datetime import datetime  
# → 현재 날짜와 시간을 기록하기 위해 datetime 모듈의 datetime 클래스 사용
import os  
# → 파일 경로를 다루기 위해 필요한 기본 OS 관련 모듈
import os  
# → 중복 import지만 문제 없음 (코드 실행에는 영향 X)
from app.core.utils import hash_password, verify_password

EXCEL_PATH = "C:\\pro\\data_log.xlsx" 
# → 현재 파이썬 파일(loge_excel.py)이 위치한 폴더 경로에 "data_log.xlsx"를 붙임
#   즉, 항상 같은 폴더 안에 있는 엑셀 파일을 자동으로 찾거나 생성하게 됨
# 1. 엑셀 초기화
def init_excel():  
    # → 프로그램 시작 시 실행됨. 
    #   엑셀 파일이 없으면 자동 생성하고, 시트(Users, LoginHistory, Comments)가 없으면 만들어줌

    if not os.path.exists(EXCEL_PATH):  
        # → data_log.xlsx 파일이 존재하지 않으면 새로 만든다
        with pd.ExcelWriter(EXCEL_PATH) as writer:  
            # → 새 엑셀 파일 생성 (pandas의 ExcelWriter 사용)
            pd.DataFrame(columns=["id", "username", "email", "password", "created_at"]).to_excel(writer, sheet_name="Users", index=False)  
            # → Users 시트 생성: 회원 정보(id, 이름, 이메일, 비밀번호, 가입일)
            pd.DataFrame(columns=["id", "user_id", "login_time"]).to_excel(writer, sheet_name="LoginHistory", index=False)  
            # → LoginHistory 시트 생성: 로그인 기록(id, 사용자id, 로그인 시간)
            pd.DataFrame(columns=["id", "post_id", "user_id", "content", "create_date"]).to_excel(writer, sheet_name="Comments", index=False)  
            # → Comments 시트 생성: 댓글 기록(id, 게시글id, 작성자id, 내용, 작성일)
    wb = load_workbook(EXCEL_PATH)  
    # → 이미 존재하는 엑셀 파일을 열어서 시트 이름을 확인함
    sheets = wb.sheetnames  
    # → 현재 엑셀 파일에 들어있는 시트들의 이름 목록을 가져옴
    wb.close()  
    # → 엑셀 파일 닫기 (다른 함수에서 다시 열 수 있도록)
    if "Users" not in sheets:  
        # → Users 시트가 없으면 새로 생성
        pd.DataFrame(columns=["id", "username", "email", "password", "created_at"]).to_excel(EXCEL_PATH, sheet_name="Users", index=False)
    if "LoginHistory" not in sheets:  
        # → LoginHistory 시트가 없으면 새로 생성
        pd.DataFrame(columns=["id", "user_id", "login_time"]).to_excel(EXCEL_PATH, sheet_name="LoginHistory", index=False)
    if "Comments" not in sheets:  
        # → Comments 시트가 없으면 새로 생성
        pd.DataFrame(columns=["id", "post_id", "user_id", "content", "create_date"]).to_excel(EXCEL_PATH, sheet_name="Comments", index=False)
    print("엑셀 시트 업데이트 완료")  
    # → 시트 확인 및 생성이 끝나면 콘솔에 메시지 출력

# 2. 사용자 등록
def register_user(username, email, password):  
    # → 회원가입 시 실행되는 함수. 새 사용자를 Users 시트에 추가함
    df = pd.read_excel(EXCEL_PATH, sheet_name="Users")  
    # → 엑셀의 Users 시트를 읽어서 DataFrame 형태로 저장
    new_id = df['id'].max() + 1 if not df.empty else 1  
    # → 기존 데이터가 있으면 마지막 id값 +1을 새 ID로 사용
    #   데이터가 없으면 첫 번째 유저는 id = 1로 시작
    df = pd.concat([df, pd.DataFrame([{
        "id": new_id,
        "username": username,
        "email": email,
        "password":  hash_password(password),
        "created_at": datetime.now()
    }])], ignore_index=True)  
    # → 새로운 사용자 정보를 기존 데이터프레임에 한 줄 추가함. 
    #   created_at은 현재 시간으로 자동 기록됨
    with pd.ExcelWriter(EXCEL_PATH, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:  
        # → 엑셀 쓰기 모드. (a: append지만, 시트 교체 옵션 사용)
        df.to_excel(writer, sheet_name="Users", index=False)  
        # → Users 시트를 새로운 DataFrame 내용으로 교체 저장
    print(f"사용자 등록 완료: {username}")  
    # → 등록 성공 메시지 출력.

# 3. 로그인 기록 저장
def save_login_history(user_id):  
    # → 사용자가 로그인할 때마다 기록을 남기는 함수
    df = pd.read_excel(EXCEL_PATH, sheet_name="LoginHistory")  
    # → LoginHistory 시트를 읽음
    new_id = df['id'].max() + 1 if not df.empty else 1  
    # → 새 기록의 ID를 설정. (기존이 있으면 +1)
    df = pd.concat([df, pd.DataFrame([{
        "id": new_id,
        "user_id": user_id,
        "login_time": datetime.now()
    }])], ignore_index=True)  
    # → 새로운 로그인 기록 한 줄 추가
    with pd.ExcelWriter(EXCEL_PATH, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:  
        # → 시트를 교체 저장
        df.to_excel(writer, sheet_name="LoginHistory", index=False)
    print(f"로그인 기록 저장 완료: user_id={user_id}")  
    # → 콘솔에 완료 메시지 출력

# 4. 댓글 추가
def add_comment(post_id, user_id, content):  
    # → 게시물에 댓글을 추가하는 함수
    df = pd.read_excel(EXCEL_PATH, sheet_name="Comments")  
    # → Comments 시트를 읽어서 DataFrame으로 저장
    new_id = df['id'].max() + 1 if not df.empty else 1  
    # → 새 댓글의 ID를 생성
    df = pd.concat([df, pd.DataFrame([{
        "id": new_id,
        "post_id": post_id,
        "user_id": user_id,
        "content": content,
        "create_date": datetime.now()
    }])], ignore_index=True)  
    # → 새로운 댓글을 DataFrame에 추가
    #   작성 시각은 datetime.now()로 기록됨
    with pd.ExcelWriter(EXCEL_PATH, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:  
        # → 엑셀에 시트를 교체 저쟝
        df.to_excel(writer, sheet_name="Comments", index=False)
    print(f"댓글 추가 완료: post_id={post_id}, user_id={user_id}")  
    # → 댓글 작성 완료 메시지 출력

# 5. 사용자 삭제 (안전한 삭제)
def delete_user_safe(username: str, email: str, password: str):  
    # → 회원 삭제 함수. 이름, 이메일, 비밀번호가 모두 일치해야 삭제 가능  
    #   단순 ID 입력만으로 다른 계정을 지우는 걸 방지함
    df_users = pd.read_excel(EXCEL_PATH, sheet_name="Users")  
    # → Users 시트 읽기
    user_row = df_users[(df_users['username'] == username) & (df_users['email'] == email)]  
    # → 입력한 이름과 이메일이 모두 일치하는 사용자 행을 찾음
    if user_row.empty:  
        # → 조건에 맞는 유저가 없을 경우
        raise ValueError("사용자가 존재하지 않습니다.")  
        # → 예외 발생시켜 API 쪽에서 에러 메시지를 보여줄 수 있음
    if user_row.iloc[0]['password'] != password:  
        # → 비밀번호가 다를 경우
        raise ValueError("비밀번호가 일치하지 않습니다.")  
        # → 비밀번호 불일치 예외 발생
    user_id = user_row.iloc[0]['id']  
    # → 삭제할 사용자의 고유 ID를 가져옴
    df_users = df_users[df_users['id'] != user_id]  
    # → Users 시트에서 해당 ID를 가진 유저 제거
    df_login = pd.read_excel(EXCEL_PATH, sheet_name="LoginHistory")  
    # → LoginHistory 시트 읽
    df_login = df_login[df_login['user_id'] != user_id]  
    # → 해당 유저의 로그인 기록 삭제
    df_comments = pd.read_excel(EXCEL_PATH, sheet_name="Comments")  
    # → Comments 시트 읽기
    df_comments = df_comments[df_comments['user_id'] != user_id]  
    # → 해당 유저의 댓글 모두 삭제
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:  
        # → 엑셀 저장 (각 시트를 교체하면서 전체 저장)
        df_users.to_excel(writer, sheet_name="Users", index=False)
        df_login.to_excel(writer, sheet_name="LoginHistory", index=False)
        df_comments.to_excel(writer, sheet_name="Comments", index=False)
    print(f"계정 삭제 완료: {username}")  
    # → 콘솔에 삭제 완료 메시지 출력

# 6. 초기 실행
init_excel()  
# → 프로그램 실행 시 자동으로 엑셀을 초기화
#   시트가 없으면 새로 만들고, 이미 있으면 그대로 둠
